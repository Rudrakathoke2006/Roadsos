import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def export_issuances_to_excel(issuances) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Issuances Register"

    # Fonts
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)

    # Colors (Deep Crimson/Slate for active university accent)
    accent_fill = PatternFill(start_color="800020", end_color="800020", fill_type="solid") # Maroon
    header_fill = PatternFill(start_color="3B3F46", end_color="3B3F46", fill_type="solid") # Dark Gray

    headers = [
        "Issuance ID", 
        "Student Name", 
        "Roll Number", 
        "Branch", 
        "Equipment Name", 
        "Quantity", 
        "Issued By Admin", 
        "Issued At", 
        "Return Due At", 
        "Returned At", 
        "Status"
    ]

    # Row 1: Title Header
    ws.merge_cells("A1:K1")
    ws["A1"] = "RBU Sports Facility Usage & Equipment Issuance Register"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = accent_fill
    ws.row_dimensions[1].height = 45

    # Row 2: Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25

    # Rows 3+: Data rows
    for row_num, item in enumerate(issuances, 3):
        student_name = item.student.name if item.student else "N/A"
        roll_number = item.student.roll_number if item.student else "N/A"
        branch = item.student.branch if item.student else "N/A"
        eq_name = item.equipment.name if item.equipment else "N/A"
        issued_by = item.issued_by_user.name if item.issued_by_user else "System"

        issued_str = item.issued_at.strftime("%Y-%m-%d %H:%M:%S") if item.issued_at else ""
        due_str = item.return_due_at.strftime("%Y-%m-%d %H:%M:%S") if item.return_due_at else ""
        returned_str = item.returned_at.strftime("%Y-%m-%d %H:%M:%S") if item.returned_at else "Active / Borrowed"

        ws.cell(row=row_num, column=1, value=str(item.id)).font = cell_font
        ws.cell(row=row_num, column=2, value=student_name).font = cell_font
        ws.cell(row=row_num, column=3, value=roll_number).font = cell_font
        ws.cell(row=row_num, column=4, value=branch).font = cell_font
        ws.cell(row=row_num, column=5, value=eq_name).font = cell_font
        ws.cell(row=row_num, column=6, value=item.quantity).font = cell_font
        ws.cell(row=row_num, column=7, value=issued_by).font = cell_font
        ws.cell(row=row_num, column=8, value=issued_str).font = cell_font
        ws.cell(row=row_num, column=9, value=due_str).font = cell_font
        ws.cell(row=row_num, column=10, value=returned_str).font = cell_font
        ws.cell(row=row_num, column=11, value=item.status).font = cell_font

        # Center numeric values or IDs, align text normally
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=11).alignment = Alignment(horizontal="center")
        ws.row_dimensions[row_num].height = 20

    # Auto-adjust column widths based on maximum contents
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            if cell.row == 1: # Ignore title block in calculation
                continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save to in-memory byte buffer
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
